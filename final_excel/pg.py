import sys
import sqlite3
from openpyxl import load_workbook

# Connect to SQLite database
conn = sqlite3.connect('books.db')
cursor = conn.cursor()

# Create tables if they don't exist
cursor.execute('''CREATE TABLE IF NOT EXISTS Cupboard (
                    id INTEGER PRIMARY KEY,
                    cupboard_number INTEGER
                )''')

cursor.execute('''CREATE TABLE IF NOT EXISTS Row (
                    id INTEGER PRIMARY KEY,
                    cupboard_id INTEGER,
                    row_number INTEGER,
                    category TEXT,
                    FOREIGN KEY (cupboard_id) REFERENCES Cupboard(id)
                )''')

cursor.execute('''CREATE TABLE IF NOT EXISTS Book (
                    id INTEGER PRIMARY KEY,
                    isbn TEXT,
                    title TEXT,
                    author TEXT,
                    publisher TEXT,
                    row_id INTEGER,
                    cupboard_id INTEGER,
                    manual_added INTEGER DEFAULT 0,
                    FOREIGN KEY (row_id, cupboard_id) REFERENCES Row(id, cupboard_id),
                    FOREIGN KEY (cupboard_id) REFERENCES Cupboard(id)
                )''')

# Function to ensure necessary cupboards and rows exist
def ensure_cupboards_and_rows_exist():
    # Ensure at least 9 cupboards exist
    cursor.execute("SELECT COUNT(*) FROM Cupboard")
    count = cursor.fetchone()[0]
    for cupboard_number in range(1, 10):
        cursor.execute("INSERT OR IGNORE INTO Cupboard (cupboard_number) VALUES (?)", (cupboard_number,))
        if count == 0:  # Only insert rows if cupboards were just created
            for row_number in range(1, 11):  # Assuming each cupboard has 10 rows
                cursor.execute(
                    "INSERT OR IGNORE INTO Row (cupboard_id, row_number) SELECT id, ? FROM Cupboard WHERE cupboard_number = ?",
                    (row_number, cupboard_number))

# Ensure necessary cupboards and rows exist
ensure_cupboards_and_rows_exist()

# Function to read data from Excel and insert into the database
def insert_data(cupboard_number, excel_filename, excel_e_filename):
    def insert_books(ws, cupboard_number, error_log):
        column_names = next(
            ws.iter_rows(min_row=1, max_row=1, values_only=True))  # Extract column names from the first row
        if column_names is None:
            print(f"Error: No column names found in {excel_filename}.")
            sys.exit(1)
        column_names_lower = [name.lower() for name in column_names if
                              name is not None]  # Convert column names to lowercase
        print(f"Processing {excel_filename}:")
        try:
            isbn_index = column_names_lower.index("isbn")
            title_index = column_names_lower.index("title")
            author_index = column_names_lower.index("author")
            publisher_index = column_names_lower.index("publisher")
            rowno_index = column_names_lower.index("row")
        except ValueError as e:
            print(f"Error: {e}")
            sys.exit(1)

        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            try:
                isbn, title, author, publisher, rowno = row[isbn_index], row[title_index], row[
                    author_index], row[publisher_index], row[rowno_index]
                if isbn is not None and isinstance(isbn, str) and isbn.strip():  # Check if ISBN is not null or empty
                    cursor.execute("SELECT id FROM Cupboard WHERE cupboard_number = ?", (cupboard_number,))
                    cupboard_row = cursor.fetchone()
                    if cupboard_row:
                        cupboard_id = cupboard_row[0]
                        cursor.execute("INSERT INTO Book (isbn, title, author, publisher, row_id, cupboard_id) VALUES (?, ?, ?, ?, ?, ?)",
                                       (isbn, title, author, publisher, rowno, cupboard_id))
                        print(f"  Inserted row {row_number}: ISBN={isbn}, Title={title}, Author={author}, Publisher={publisher}, ROWNO={rowno}")
            except ValueError as e:
                print(f"Error: {e}")
                error_log.append((cupboard_number, row_number, str(e)))
                continue
            except Exception as e:
                print(f"Error: {e}")
                error_log.append((cupboard_number, row_number, str(e)))
                continue

    error_log = []

    # Load workbook for cupboard.xlsx
    wb = load_workbook(excel_filename)
    ws = wb.active
    insert_books(ws, cupboard_number, error_log)

    # Load workbook for cupboardno_e.xlsx
    wb_e = load_workbook(excel_e_filename)
    ws_e = wb_e.active
    insert_books(ws_e, cupboard_number, error_log)

    if error_log:
        print(f"Errors encountered while processing {excel_filename}:")
        for error in error_log:
            print(f"  Error in row {error[1]}: {error[2]}")

# Insert data for each cupboard
for cupboard_number in range(1, 10):
    insert_data(cupboard_number, f'{cupboard_number}.xlsx', f'{cupboard_number}_e.xlsx')

# Commit changes and close connection
conn.commit()
conn.close()

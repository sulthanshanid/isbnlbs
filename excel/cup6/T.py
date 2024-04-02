import os
import requests
import json
import re
import openpyxl

# Function to process image files in a directory
def process_images_in_directory(directory):
    for filename in os.listdir(directory):
        if filename.endswith(".jpeg") or filename.endswith(".png"):
            process_image(os.path.join(directory, filename), directory)

# Function to process an image
def process_image(image_path, directory_name):
    # First Request - Upload Image
    url_upload = 'https://www.blackbox.ai/api/upload'
    files = {'image': open(image_path, 'rb')}
    headers_upload = {
        # Add your headers here
    }
    response_upload = requests.post(url_upload, files=files, headers=headers_upload)

    # Ensure the upload was successful
    if response_upload.status_code != 200:
        print("Error uploading image:", response_upload.status_code)
        return

    # Convert the response content from bytes to string
    response_content_str = response_upload.content.decode('utf-8')

    # Second Request - Chat API with First Response
    url_chat = 'https://www.blackbox.ai/api/chat'

    payload = {
        "messages": [
            {
                "id": "6ZUkhHE",
                "content": response_content_str + "give me answer in json format extract title ,author,publisher ",
                # Assuming response is JSON, adjust accordingly
                "role": "user"
            }
        ],
        "id": "6ZUkhHE",
        "previewToken": None,
        "userId": "64ef24a9a502c400325b1793",
        "codeModelMode": False,
        "agentMode": {},
        "trendingAgentMode": {},
        "isMicMode": False,
        "isChromeExt": False,
        "githubToken": None
    }

    headers_chat = {
        # Add your headers here
    }

    response_chat = requests.post(url_chat, data=json.dumps(payload), headers=headers_chat)

    # Ensure the chat request was successful
    if response_chat.status_code != 200:
        print("Error with chat request:", response_chat.status_code)
        return

    # Extract JSON data from chat response
    json_pattern = r'{(?:[^{}]|{(?:[^{}]|{[^{}]})})*}'

    json_match = re.search(json_pattern, response_chat.text)

    if json_match:
        # Extract the JSON string
        json_string = json_match.group(0)

        # Parse the JSON string
        try:
            chat_data = json.loads(json_string)
            title = chat_data.get('title')
            authors = chat_data.get('author')
            if authors:
                if isinstance(authors, list):
                    author = authors[0]  # Select the first author if available
                else:
                    # Split the string of authors by commas and take the first author
                    author = authors.split(',')[0].strip()
            else:
                author = "Unknown"  # Set a default value if no author is provided
            publisher = chat_data.get('publisher')
            print("Title:", title)
            print("Author:", author)
            print("Publisher:", publisher)
        except json.JSONDecodeError as e:
            print("Error decoding JSON:", e)
            return
        isbn = None
        # Third Request - OpenLibrary Search
        url_search = f"https://openlibrary.org/search?q={title}+{author}"
        headers_search = {
            # Add your headers here
        }

        response_search = requests.get(url_search, headers=headers_search)

        if response_search.status_code == 200:
            try:
                response_json = json.loads(response_search.content)

                # Extract ISBNs
                isbn_list = []
                for doc in response_json['docs']:
                    if 'isbn' in doc:
                        isbn_list.extend(doc['isbn'])

                # If no ISBNs found, set ISBN to None
                if not isbn_list:
                    isbn = None
                else:
                    # Get the first ISBN
                    isbn = isbn_list[0]

                # Print the ISBN if available
                if isbn:
                    print("ISBN:", isbn)
                    # Save entry to Excel file

                else:
                    print("ISBN not found")

            except Exception as e:
                print(e)

        else:
            print("Error with OpenLibrary search:", response_search.status_code)
        save_to_excel(title, author, publisher, isbn, directory_name)

    else:
        print("No JSON found in the response string.")

# Function to save the extracted information to an Excel file
def save_to_excel(title, author, publisher, isbn, directory_name):
    # Get current working directory
    cwd = os.getcwd()

    # Create or load workbook
    excel_file_path = os.path.join(cwd, '3manualadd.xlsx')
    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
    else:
        workbook = openpyxl.Workbook()

    # Get or create worksheet
    worksheet_name = "Sheet1"
    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(title=worksheet_name)

    # Find the next empty row
    next_row = len(worksheet["A"]) + 1

    # Write data to cells
    worksheet.cell(row=next_row, column=1, value=title)
    worksheet.cell(row=next_row, column=2, value=author)
    worksheet.cell(row=next_row, column=3, value=publisher)
    worksheet.cell(row=next_row, column=4, value=isbn)
    worksheet.cell(row=next_row, column=5, value=directory_name)  # Directory name in "ROW" column

    # Save workbook
    workbook.save(excel_file_path)

# Iterate over each directory in the working directory
for directory in os.listdir():
    if os.path.isdir(directory):
        process_images_in_directory(directory)

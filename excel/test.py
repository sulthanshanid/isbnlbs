import sys
import sqlite3
from openpyxl import load_workbook

# Connect to SQLite database
conn = sqlite3.connect('books.db')
cursor = conn.cursor()

# Create tables if they don't exist
cursor.execute('''CREATE TABLE IF NOT EXISTS Cupboard (
                    id INTEGER PRIMARY KEY,
                    cupboard_number INTEGER
                )''')

cursor.execute('''CREATE TABLE IF NOT EXISTS Row (
                    id INTEGER PRIMARY KEY,
                    cupboard_id INTEGER,
                    row_number INTEGER,
                    category TEXT,
                    FOREIGN KEY (cupboard_id) REFERENCES Cupboard(id)
                )''')

cursor.execute('''CREATE TABLE IF NOT EXISTS Book (
                    id INTEGER PRIMARY KEY,
                    isbn TEXT,
                    title TEXT,
                    author TEXT,
                    publisher TEXT,
                    row_id INTEGER,
                    cupboard_id INTEGER,
                    manual_added INTEGER DEFAULT 0, -- New column to mark manually added books
                    FOREIGN KEY (row_id) REFERENCES Row(id),
                    FOREIGN KEY (cupboard_id) REFERENCES Cupboard(id)
                )''')


# Function to ensure necessary cupboards and rows exist
def ensure_cupboards_and_rows_exist():
    # Ensure at least 9 cupboards exist
    cursor.execute("SELECT COUNT(*) FROM Cupboard")
    count = cursor.fetchone()[0]
    for cupboard_number in range(1, 10):
        cursor.execute("INSERT OR IGNORE INTO Cupboard (cupboard_number) VALUES (?)", (cupboard_number,))
        if count == 0:  # Only insert rows if cupboards were just created
            for row_number in range(1, 11):  # Assuming each cupboard has 10 rows
                cursor.execute(
                    "INSERT OR IGNORE INTO Row (cupboard_id, row_number) SELECT id, ? FROM Cupboard WHERE cupboard_number = ?",
                    (row_number, cupboard_number))


# Ensure necessary cupboards and rows exist
ensure_cupboards_and_rows_exist()


# Function to read data from Excel and insert into the database
def insert_data(cupboard_number, excel_filename, excel_e_filename):
    def insert_books(ws, cupboard_number, error_log):
        column_names = next(
            ws.iter_rows(min_row=1, max_row=1, values_only=True))  # Extract column names from the first row
        column_names_lower = [name.lower() for name in column_names]  # Convert column names to lowercase
        print("Column Names:", column_names)  # Print out the extracted column names
        try:
            isbn_index = column_names_lower.index("isbn")
            title_index = column_names_lower.index("title")
            author_index = column_names_lower.index("author")
            publisher_index = column_names_lower.index("publisher")
            row_index = column_names_lower.index("row")
        except ValueError as e:
            print(f"Error in {excel_filename}: {e}")
            sys.exit(1)

        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                isbn, title, author, publisher, original_row_number = row[isbn_index], row[title_index], row[
                    author_index], row[publisher_index], row[row_index]
                if isbn is not None and isinstance(isbn, str) and isbn.strip():  # Check if ISBN is not null or empty
                    cursor.execute("SELECT id FROM Cupboard WHERE cupboard_number = ?", (cupboard_number,))
                    cupboard_row = cursor.fetchone()
                    if cupboard_row:
                        cupboard_id = cupboard_row[0]
                        cursor.execute("SELECT id FROM Row WHERE cupboard_id = ? AND row_number = ?",
                                       (cupboard_id, original_row_number))
                        row_row = cursor.fetchone()
                        if row_row:
                            row_id = row_row[0]
                            cursor.execute(
                                "INSERT INTO Book (isbn, title, author, publisher, row_id, cupboard_id) VALUES (?, ?, ?, ?, ?, ?)",
                                (isbn, title, author, publisher, row_id, cupboard_id))
            except ValueError as e:
                print(f"Error in {excel_filename}: {e}")
                sys.exit(1)
            except Exception as e:
                error_log.append((excel_filename, original_row_number, str(e)))
                continue

    error_log = []

    # Load workbook for cuboard.xlsx
    wb = load_workbook(excel_filename)
    ws = wb.active
    insert_books(ws, cupboard_number, error_log)

    # Load workbook for cuboardno_e.xlsx
    wb_e = load_workbook(excel_e_filename)
    ws_e = wb_e.active
    insert_books(ws_e, cupboard_number, error_log)

    # Print any errors
    for error in error_log:
        print(f"Error in {error[0]}, original row {error[1]}: {error[2]}")

    if error_log:  # Exit with error code if there are errors
        sys.exit(1)

    conn.commit()


# Insert data for each cupboard
for cupboard_number in range(1, 10):
    insert_data(cupboard_number, f'{cupboard_number}.xlsx', f'{cupboard_number}_e.xlsx')

# Close connection
conn.close()
